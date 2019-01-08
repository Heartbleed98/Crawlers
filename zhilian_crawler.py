import importlib,sys 
import json, pprint
import time
import requests
import numpy as np
from openpyxl import Workbook

importlib.reload(sys)

def crawler_main():
	start = 60
	job_list = []
	count = 1

	url1 = 'https://fe-api.zhaopin.com/c/i/sou?pageSize=60&cityId=489&workExperience=-1&education=-1&companyType=-1&employmentType=-1&jobWelfareTag=-1&jobType=2040&kt=3&lastUrlQuery=%7B%22jl%22:%22489%22,%22jt%22:%2223,160000,2040%22%7D'
	r = requests.get(url1)
	r.encoding = 'utf-8'
	r_json = r.json()
	r_json = r_json['data']['results']	

	for i in range(len(r_json)):
		city = r_json[i]['city']['display'].split('-')
		city = city[0]

		company = r_json[i]['company']['name']
		job = r_json[i]['jobName']
		eduLevel = r_json[i]['eduLevel']['name']
		salary = r_json[i]['salary']
		workingExp = r_json[i]['workingExp']['name']
		job_list.append([city,company,job,eduLevel,salary,workingExp])
		
	print('Getting information on page 1.')

	while (count <= 20):
		time.sleep(np.random.rand()*5)
		url2 = 'https://fe-api.zhaopin.com/c/i/sou?start=' + str(start) + '&pageSize=60&cityId=489&workExperience=-1&education=-1&companyType=-1&employmentType=-1&jobWelfareTag=-1&jobType=2040&kt=3&lastUrlQuery=%7B%22p%22:2,%22jl%22:%22489%22,%22jt%22:%2223,160000,2040%22%7D'
		r = requests.get(url2)
		r.encoding = 'utf-8'
		r_json = r.json()
		r_json = r_json['data']['results']	

		for i in range(len(r_json)):
			city = r_json[i]['city']['display'].split('-')
			city = city[0]
			
			company = r_json[i]['company']['name']
			job = r_json[i]['jobName']
			eduLevel = r_json[i]['eduLevel']['name']
			salary = r_json[i]['salary']
			workingExp = r_json[i]['workingExp']['name']
			job_list.append([city,company,job,eduLevel,salary,workingExp])

		start += 60
		count += 1

		print('Getting information.')

	excel_output(job_list)


def excel_output(job_list):
    wb=Workbook()
    ws=[]
    ws.append(wb.create_sheet(title='智联招聘数据')) #utf8->unicode

    ws[0].append(['城市','公司','招聘职位','学历要求','月薪','工作经验'])
    for jl in job_list:
        ws[0].append([jl[0],jl[1],jl[2],jl[3],jl[4],jl[5]])
    save_path='智联招聘数据'
    save_path+='.xlsx'
    wb.save(save_path)

if __name__=='__main__':
	crawler_main()